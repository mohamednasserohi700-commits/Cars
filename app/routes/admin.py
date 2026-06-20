from flask import Blueprint, render_template, redirect, url_for, request, flash, jsonify, send_file
from flask_login import login_required, current_user
from datetime import datetime, date
from sqlalchemy import func
import io
import json
import openpyxl
from app import db
from app.models import Employee, Bus, Station, Driver, Registration, LoginLog, Settings, Admin

admin_bp = Blueprint('admin', __name__)


def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated


# ─── DASHBOARD ────────────────────────────────────────────────────────────────

@admin_bp.route('/dashboard')
@login_required
def dashboard():
    today = date.today()
    stats = {
        'employees': Employee.query.filter_by(is_active=True).count(),
        'buses': Bus.query.filter_by(is_active=True).count(),
        'drivers': Driver.query.count(),
        'total_registrations': Registration.query.count(),
        'today_registrations': Registration.query.filter(
            func.date(Registration.registration_date) == today).count(),
        'full_buses': sum(1 for b in Bus.query.filter_by(is_active=True).all() if b.is_full()),
    }
    # Shift stats for today
    shifts = db.session.query(
        Registration.shift, func.count(Registration.id)
    ).filter(func.date(Registration.registration_date) == today).group_by(Registration.shift).all()
    shift_data = {s[0]: s[1] for s in shifts}

    # Last 7 days registrations
    from sqlalchemy import text
    daily = db.session.execute(
        text("SELECT date(registration_date) as d, count(*) as c FROM registrations GROUP BY d ORDER BY d DESC LIMIT 7")
    ).fetchall()
    daily_labels = [str(r[0]) for r in reversed(daily)]
    daily_values = [r[1] for r in reversed(daily)]

    return render_template('admin/dashboard.html', stats=stats, shift_data=shift_data,
                           daily_labels=daily_labels, daily_values=daily_values)


# ─── EMPLOYEES ────────────────────────────────────────────────────────────────

@admin_bp.route('/employees')
@login_required
def employees():
    q = request.args.get('q', '')
    query = Employee.query
    if q:
        query = query.filter(
            (Employee.name.ilike(f'%{q}%')) | (Employee.global_id.ilike(f'%{q}%')) | (Registration.guest_global_id.ilike(f'%{q}%'))
        )
    employees = query.order_by(Employee.created_at.desc()).all()
    return render_template('admin/employees/index.html', employees=employees, q=q)


@admin_bp.route('/employees/add', methods=['GET', 'POST'])
@login_required
def add_employee():
    if request.method == 'POST':
        global_id = request.form.get('global_id', '').strip()
        name = request.form.get('name', '').strip()
        department = request.form.get('department', '').strip()
        affiliate = request.form.get('affiliate', 'مكتب')

        if not global_id or not name:
            flash('الرقم العالمي والاسم مطلوبان.', 'danger')
            return redirect(url_for('admin.add_employee'))

        if Employee.query.filter_by(global_id=global_id).first():
            flash('الرقم العالمي موجود مسبقاً.', 'warning')
            return redirect(url_for('admin.add_employee'))

        emp = Employee(global_id=global_id, name=name, department=department, affiliate=affiliate)
        db.session.add(emp)
        db.session.commit()
        flash('تم إضافة الموظف بنجاح.', 'success')
        return redirect(url_for('admin.employees'))
    return render_template('admin/employees/form.html', employee=None)


@admin_bp.route('/employees/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_employee(id):
    emp = Employee.query.get_or_404(id)
    if request.method == 'POST':
        emp.name = request.form.get('name', '').strip()
        emp.department = request.form.get('department', '').strip()
        emp.affiliate = request.form.get('affiliate', 'مكتب')
        emp.is_active = 'is_active' in request.form
        db.session.commit()
        flash('تم تحديث بيانات الموظف.', 'success')
        return redirect(url_for('admin.employees'))
    return render_template('admin/employees/form.html', employee=emp)


@admin_bp.route('/employees/<int:id>/delete', methods=['POST'])
@login_required
def delete_employee(id):
    emp = Employee.query.get_or_404(id)
    db.session.delete(emp)
    db.session.commit()
    flash('تم حذف الموظف.', 'success')
    return redirect(url_for('admin.employees'))


@admin_bp.route('/employees/import', methods=['POST'])
@login_required
def import_employees():
    file = request.files.get('file')
    if not file:
        flash('لم يتم اختيار ملف.', 'danger')
        return redirect(url_for('admin.employees'))
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            global_id = str(row[0]).strip()
            name = str(row[1]).strip() if row[1] else ''
            dept = str(row[2]).strip() if row[2] else ''
            affiliate = str(row[3]).strip() if len(row) > 3 and row[3] else 'مكتب'
            if not Employee.query.filter_by(global_id=global_id).first():
                db.session.add(Employee(global_id=global_id, name=name, department=dept, affiliate=affiliate))
                count += 1
        db.session.commit()
        flash(f'تم استيراد {count} موظف بنجاح.', 'success')
    except Exception as e:
        flash(f'خطأ في الاستيراد: {str(e)}', 'danger')
    return redirect(url_for('admin.employees'))


@admin_bp.route('/employees/export')
@login_required
def export_employees():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'الموظفون'
    ws.append(['الرقم العالمي', 'الاسم', 'القسم', 'التابع', 'الحالة'])
    for emp in Employee.query.all():
        ws.append([emp.global_id, emp.name, emp.department, emp.affiliate,
                   'نشط' if emp.is_active else 'غير نشط'])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='employees.xlsx')


# ─── BUSES ────────────────────────────────────────────────────────────────────

@admin_bp.route('/buses/<int:id>/detail')
@login_required
def bus_detail(id):
    from flask import render_template
    bus = Bus.query.get_or_404(id)
    today = date.today()
    today_regs = Registration.query.filter_by(bus_id=id).filter(
        func.date(Registration.registration_date) == today
    ).all()
    all_regs = Registration.query.filter_by(bus_id=id).order_by(
        Registration.registration_date.desc()).limit(50).all()
    return render_template('admin/buses/detail.html', bus=bus,
                           today_regs=today_regs, all_regs=all_regs)


@admin_bp.route('/buses')
@login_required
def buses():
    buses = Bus.query.all()
    return render_template('admin/buses/index.html', buses=buses)


@admin_bp.route('/buses/add', methods=['GET', 'POST'])
@login_required
def add_bus():
    drivers = Driver.query.all()
    all_buses = Bus.query.all()
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        total_seats = int(request.form.get('total_seats', 50))
        driver_id = request.form.get('driver_id') or None
        backup_driver_id = request.form.get('backup_driver_id') or None
        backup_bus_id = request.form.get('backup_bus_id') or None

        bus = Bus(name=name, total_seats=total_seats,
                  driver_id=driver_id, backup_driver_id=backup_driver_id,
                  backup_bus_id=backup_bus_id)
        db.session.add(bus)
        db.session.commit()
        flash('تم إضافة الخط بنجاح.', 'success')
        return redirect(url_for('admin.buses'))
    return render_template('admin/buses/form.html', bus=None, drivers=drivers, all_buses=all_buses)


@admin_bp.route('/buses/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_bus(id):
    bus = Bus.query.get_or_404(id)
    drivers = Driver.query.all()
    all_buses = Bus.query.filter(Bus.id != id).all()
    if request.method == 'POST':
        bus.name = request.form.get('name', '').strip()
        bus.total_seats = int(request.form.get('total_seats', 50))
        bus.driver_id = request.form.get('driver_id') or None
        bus.backup_driver_id = request.form.get('backup_driver_id') or None
        bus.backup_bus_id = request.form.get('backup_bus_id') or None
        bus.is_active = 'is_active' in request.form
        db.session.commit()
        flash('تم تحديث الخط.', 'success')
        return redirect(url_for('admin.buses'))
    return render_template('admin/buses/form.html', bus=bus, drivers=drivers, all_buses=all_buses)


@admin_bp.route('/buses/<int:id>/delete', methods=['POST'])
@login_required
def delete_bus(id):
    bus = Bus.query.get_or_404(id)
    if bus.registrations.count() > 0:
        flash('لا يمكن حذف خط يحتوي على تسجيلات.', 'danger')
        return redirect(url_for('admin.buses'))
    db.session.delete(bus)
    db.session.commit()
    flash('تم حذف الخط.', 'success')
    return redirect(url_for('admin.buses'))


# ─── STATIONS ────────────────────────────────────────────────────────────────

@admin_bp.route('/buses/<int:bus_id>/stations')
@login_required
def stations(bus_id):
    bus = Bus.query.get_or_404(bus_id)
    stations = Station.query.filter_by(bus_id=bus_id).order_by(Station.order).all()
    return render_template('admin/stations/index.html', bus=bus, stations=stations)


@admin_bp.route('/buses/<int:bus_id>/stations/add', methods=['POST'])
@login_required
def add_station(bus_id):
    name = request.form.get('name', '').strip()
    order = int(request.form.get('order', 0))
    if name:
        db.session.add(Station(name=name, bus_id=bus_id, order=order))
        db.session.commit()
        flash('تم إضافة المحطة.', 'success')
    return redirect(url_for('admin.stations', bus_id=bus_id))


@admin_bp.route('/stations/<int:id>/delete', methods=['POST'])
@login_required
def delete_station(id):
    station = Station.query.get_or_404(id)
    bus_id = station.bus_id
    db.session.delete(station)
    db.session.commit()
    flash('تم حذف المحطة.', 'success')
    return redirect(url_for('admin.stations', bus_id=bus_id))


# ─── DRIVERS ─────────────────────────────────────────────────────────────────

@admin_bp.route('/drivers')
@login_required
def drivers():
    drivers = Driver.query.all()
    return render_template('admin/drivers/index.html', drivers=drivers)


@admin_bp.route('/drivers/add', methods=['GET', 'POST'])
@login_required
def add_driver():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        phone = request.form.get('phone', '').strip()
        if name and phone:
            db.session.add(Driver(name=name, phone=phone))
            db.session.commit()
            flash('تم إضافة السائق.', 'success')
            return redirect(url_for('admin.drivers'))
    return render_template('admin/drivers/form.html', driver=None)


@admin_bp.route('/drivers/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_driver(id):
    driver = Driver.query.get_or_404(id)
    if request.method == 'POST':
        driver.name = request.form.get('name', '').strip()
        driver.phone = request.form.get('phone', '').strip()
        db.session.commit()
        flash('تم تحديث بيانات السائق.', 'success')
        return redirect(url_for('admin.drivers'))
    return render_template('admin/drivers/form.html', driver=driver)


@admin_bp.route('/drivers/<int:id>/delete', methods=['POST'])
@login_required
def delete_driver(id):
    driver = Driver.query.get_or_404(id)
    if driver.primary_buses.count() > 0 or driver.backup_buses.count() > 0:
        flash('لا يمكن حذف سائق مرتبط بخط.', 'danger')
        return redirect(url_for('admin.drivers'))
    db.session.delete(driver)
    db.session.commit()
    flash('تم حذف السائق.', 'success')
    return redirect(url_for('admin.drivers'))


# ─── REGISTRATIONS ────────────────────────────────────────────────────────────

@admin_bp.route('/registrations')
@login_required
def registrations():
    q = request.args.get('q', '')
    shift_f = request.args.get('shift', '')
    date_f = request.args.get('date', '')
    query = Registration.query.outerjoin(Employee).join(Bus)
    if q:
        query = query.filter(
            (Employee.name.ilike(f'%{q}%')) | (Employee.global_id.ilike(f'%{q}%')) | (Registration.guest_global_id.ilike(f'%{q}%'))
        )
    if shift_f:
        query = query.filter(Registration.shift == shift_f)
    if date_f:
        try:
            d = datetime.strptime(date_f, '%Y-%m-%d').date()
            query = query.filter(func.date(Registration.registration_date) == d)
        except:
            pass
    regs = query.order_by(Registration.registration_date.desc()).all()

    # Count registrations per bus line, based on the same filtered results shown in the table
    bus_counts = {}
    for r in regs:
        bus_counts[r.bus.name] = bus_counts.get(r.bus.name, 0) + 1
    line_stats = sorted(bus_counts.items(), key=lambda x: x[1], reverse=True)

    return render_template('admin/registrations/index.html', registrations=regs, q=q,
                           shift_f=shift_f, date_f=date_f, line_stats=line_stats)


@admin_bp.route('/registrations/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_registration(id):
    from flask import render_template
    reg = Registration.query.get_or_404(id)
    buses = Bus.query.filter_by(is_active=True).all()
    if request.method == 'POST':
        bus_id = request.form.get('bus_id')
        station_id = request.form.get('station_id')
        shift = request.form.get('shift')
        phone = request.form.get('phone', '').strip()
        travel_date_str = request.form.get('travel_date')
        try:
            reg.travel_date = datetime.strptime(travel_date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('تاريخ غير صحيح.', 'danger')
            return redirect(url_for('admin.edit_registration', id=id))
        reg.bus_id = bus_id
        reg.station_id = station_id
        reg.shift = shift
        reg.phone = phone
        db.session.commit()
        flash('تم تحديث التسجيل بنجاح.', 'success')
        return redirect(url_for('admin.registrations'))
    return render_template('admin/registrations/form.html', reg=reg, buses=buses)


@admin_bp.route('/registrations/<int:id>/delete', methods=['POST'])
@login_required
def delete_registration(id):
    reg = Registration.query.get_or_404(id)
    db.session.delete(reg)
    db.session.commit()
    flash('تم حذف التسجيل.', 'success')
    return redirect(url_for('admin.registrations'))


@admin_bp.route('/registrations/delete-all', methods=['POST'])
@login_required
def delete_all_registrations():
    q = request.form.get('q', '')
    shift_f = request.form.get('shift', '')
    date_f = request.form.get('date', '')

    query = Registration.query.outerjoin(Employee).join(Bus)
    if q:
        query = query.filter(
            (Employee.name.ilike(f'%{q}%')) | (Employee.global_id.ilike(f'%{q}%')) | (Registration.guest_global_id.ilike(f'%{q}%'))
        )
    if shift_f:
        query = query.filter(Registration.shift == shift_f)
    if date_f:
        try:
            d = datetime.strptime(date_f, '%Y-%m-%d').date()
            query = query.filter(func.date(Registration.registration_date) == d)
        except:
            pass

    count = query.count()
    ids = [r.id for r in query.all()]
    if ids:
        Registration.query.filter(Registration.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()

    flash(f'تم حذف {count} تسجيل بنجاح.', 'success')
    return redirect(url_for('admin.registrations', q=q, shift=shift_f, date=date_f))


@admin_bp.route('/registrations/export')
@login_required
def export_registrations():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'التسجيلات'
    ws.append(['الاسم', 'الرقم العالمي', 'الخط', 'المحطة', 'الوردية',
               'يوم السفر', 'تاريخ التسجيل', 'وقت التسجيل', 'رقم الهاتف'])
    for r in Registration.query.join(Employee).all():
        ws.append([
            r.employee.name, r.employee.global_id, r.bus.name, r.station.name,
            r.shift, str(r.travel_date),
            r.registration_date.strftime('%Y-%m-%d'), r.registration_date.strftime('%H:%M:%S'),
            r.phone
        ])
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='registrations.xlsx')


@admin_bp.route('/registrations/export-pdf')
@login_required
def export_registrations_pdf():
    from app.services.pdf_service import generate_registrations_pdf
    regs = Registration.query.join(Employee).order_by(Registration.registration_date.desc()).all()
    pdf_buf = generate_registrations_pdf(regs)
    if pdf_buf is None:
        flash('مكتبة PDF غير متاحة. استخدم تصدير Excel.', 'warning')
        return redirect(url_for('admin.registrations'))
    return send_file(pdf_buf, mimetype='application/pdf',
                     as_attachment=True, download_name='registrations.pdf')


@admin_bp.route('/employees/export-pdf')
@login_required
def export_employees_pdf():
    from app.services.pdf_service import generate_employees_pdf
    emps = Employee.query.all()
    pdf_buf = generate_employees_pdf(emps)
    if pdf_buf is None:
        flash('مكتبة PDF غير متاحة.', 'warning')
        return redirect(url_for('admin.employees'))
    return send_file(pdf_buf, mimetype='application/pdf',
                     as_attachment=True, download_name='employees.pdf')


@admin_bp.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    from flask import render_template
    if request.method == 'POST':
        old_pw = request.form.get('old_password', '')
        new_pw = request.form.get('new_password', '')
        confirm_pw = request.form.get('confirm_password', '')
        if not current_user.check_password(old_pw):
            flash('كلمة المرور الحالية غير صحيحة.', 'danger')
        elif new_pw != confirm_pw:
            flash('كلمة المرور الجديدة غير متطابقة.', 'danger')
        elif len(new_pw) < 6:
            flash('كلمة المرور يجب أن تكون 6 أحرف على الأقل.', 'danger')
        else:
            current_user.set_password(new_pw)
            db.session.commit()
            flash('تم تغيير كلمة المرور بنجاح.', 'success')
            return redirect(url_for('admin.dashboard'))
    return render_template('admin/change_password.html')




# ─── LOGIN LOGS ───────────────────────────────────────────────────────────────

@admin_bp.route('/login-logs')
@login_required
def login_logs():
    logs = LoginLog.query.order_by(LoginLog.login_time.desc()).limit(500).all()
    return render_template('admin/logs/index.html', logs=logs)


@admin_bp.route('/login-logs/delete-all', methods=['POST'])
@login_required
def delete_all_login_logs():
    count = LoginLog.query.count()
    LoginLog.query.delete()
    db.session.commit()
    flash(f'تم حذف جميع سجلات الدخول بنجاح ({count} سجل).', 'success')
    return redirect(url_for('admin.login_logs'))


# ─── SETTINGS ────────────────────────────────────────────────────────────────

@admin_bp.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    if request.method == 'POST':
        Settings.set('registration_open', '1' if 'registration_open' in request.form else '0')
        allowed_days = request.form.getlist('allowed_days')
        Settings.set('allowed_days', ','.join(allowed_days))
        Settings.set('time_from', request.form.get('time_from', '00:00'))
        Settings.set('time_to', request.form.get('time_to', '23:59'))
        Settings.set('contact_phone', request.form.get('contact_phone', ''))
        Settings.set('company_name', request.form.get('company_name', ''))
        Settings.set('theme', request.form.get('theme', 'light'))

        # Handle logo upload
        logo = request.files.get('logo')
        if logo and logo.filename:
            import os
            from flask import current_app
            ext = logo.filename.rsplit('.', 1)[-1].lower()
            if ext in ['png', 'jpg', 'jpeg', 'gif', 'svg']:
                path = os.path.join(current_app.config['UPLOAD_FOLDER'], f'logo.{ext}')
                logo.save(path)
                Settings.set('logo', f'logo.{ext}')

        flash('تم حفظ الإعدادات بنجاح.', 'success')
        return redirect(url_for('admin.settings'))

    current_settings = {
        'registration_open': Settings.get('registration_open', '1'),
        'allowed_days': Settings.get('allowed_days', '0,1,2,3,4').split(','),
        'time_from': Settings.get('time_from', '00:00'),
        'time_to': Settings.get('time_to', '23:59'),
        'contact_phone': Settings.get('contact_phone', ''),
        'company_name': Settings.get('company_name', 'شركتنا'),
        'theme': Settings.get('theme', 'light'),
        'logo': Settings.get('logo', ''),
    }
    return render_template('admin/settings/index.html', settings=current_settings)


# ─── FULL DATABASE BACKUP (EXPORT / IMPORT) ────────────────────────────────────

# Order matters: parents before children, to respect foreign key dependencies
# on both export (for readability) and import (so FKs resolve correctly).
BACKUP_MODELS = [
    ('admins', Admin, ['id', 'username', 'password_hash', 'created_at']),
    ('drivers', Driver, ['id', 'name', 'phone', 'created_at']),
    ('employees', Employee, ['id', 'global_id', 'name', 'department', 'affiliate',
                              'is_active', 'created_at', 'updated_at']),
    ('buses', Bus, ['id', 'name', 'total_seats', 'is_active', 'driver_id',
                     'backup_driver_id', 'backup_bus_id', 'created_at']),
    ('stations', Station, ['id', 'name', 'bus_id', 'order', 'created_at']),
    ('registrations', Registration, ['id', 'employee_id', 'guest_global_id', 'employee_name',
                                      'bus_id', 'station_id', 'shift', 'phone', 'pickup_time',
                                      'travel_date', 'registration_date', 'affiliate']),
    ('login_logs', LoginLog, ['id', 'employee_id', 'global_id', 'ip_address', 'device_type',
                               'browser', 'status', 'login_time']),
    ('settings', Settings, ['id', 'key', 'value', 'updated_at']),
]

BACKUP_FORMAT_VERSION = 1


def _serialize_value(value):
    """Convert a column value into something JSON-safe, preserving type info."""
    if isinstance(value, datetime):
        return {'__type__': 'datetime', 'value': value.isoformat()}
    if isinstance(value, date):
        return {'__type__': 'date', 'value': value.isoformat()}
    return value


def _deserialize_value(value):
    """Reverse of _serialize_value."""
    if isinstance(value, dict) and '__type__' in value:
        if value['__type__'] == 'datetime':
            return datetime.fromisoformat(value['value'])
        if value['__type__'] == 'date':
            return date.fromisoformat(value['value'])
    return value


@admin_bp.route('/settings/backup/export')
@login_required
def export_full_backup():
    """Export every table in the database into a single accurate JSON file."""
    backup = {
        'format_version': BACKUP_FORMAT_VERSION,
        'exported_at': datetime.now().isoformat(),
        'tables': {}
    }

    for table_name, model, columns in BACKUP_MODELS:
        rows = []
        for record in model.query.order_by(model.id).all():
            row = {col: _serialize_value(getattr(record, col)) for col in columns}
            rows.append(row)
        backup['tables'][table_name] = rows

    json_bytes = json.dumps(backup, ensure_ascii=False, indent=2).encode('utf-8')
    buffer = io.BytesIO(json_bytes)
    buffer.seek(0)

    filename = f"backup_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.json"
    return send_file(buffer, as_attachment=True, download_name=filename,
                      mimetype='application/json')


@admin_bp.route('/settings/backup/import', methods=['POST'])
@login_required
def import_full_backup():
    """Restore the database from a backup JSON file produced by export_full_backup.

    mode = 'replace' -> wipe all tables then insert the backup data (exact restore)
    mode = 'merge'    -> keep existing data, only add rows whose primary key
                          does not already exist in each table
    """
    file = request.files.get('backup_file')
    mode = request.form.get('import_mode', 'merge')

    if not file or not file.filename:
        flash('يرجى اختيار ملف النسخة الاحتياطية.', 'danger')
        return redirect(url_for('admin.settings'))

    try:
        data = json.load(file.stream)
    except Exception:
        flash('الملف غير صالح أو تالف. يرجى اختيار ملف نسخة احتياطية صحيح (JSON).', 'danger')
        return redirect(url_for('admin.settings'))

    if 'tables' not in data:
        flash('صيغة الملف غير معروفة. تأكد أنه ملف نسخة احتياطية تم تصديره من هذا النظام.', 'danger')
        return redirect(url_for('admin.settings'))

    tables = data['tables']

    try:
        if mode == 'replace':
            # Delete in reverse dependency order to avoid FK violations
            for table_name, model, _ in reversed(BACKUP_MODELS):
                model.query.delete()
            db.session.flush()

            for table_name, model, columns in BACKUP_MODELS:
                for row in tables.get(table_name, []):
                    kwargs = {col: _deserialize_value(row.get(col)) for col in columns}
                    db.session.add(model(**kwargs))
        else:
            # Merge: only insert rows whose id doesn't already exist
            for table_name, model, columns in BACKUP_MODELS:
                existing_ids = {row_id for (row_id,) in db.session.query(model.id).all()}
                for row in tables.get(table_name, []):
                    if row.get('id') in existing_ids:
                        continue
                    kwargs = {col: _deserialize_value(row.get(col)) for col in columns}
                    db.session.add(model(**kwargs))

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء الاستيراد، تم التراجع عن أي تغييرات: {e}', 'danger')
        return redirect(url_for('admin.settings'))

    flash('تم استيراد النسخة الاحتياطية بنجاح.', 'success')
    return redirect(url_for('admin.settings'))


# ─── REPORTS ─────────────────────────────────────────────────────────────────

@admin_bp.route('/reports')
@login_required
def reports():
    # By bus
    bus_stats = db.session.query(
        Bus.name, func.count(Registration.id)
    ).join(Registration, Bus.id == Registration.bus_id).group_by(Bus.name).all()

    # By shift
    shift_stats = db.session.query(
        Registration.shift, func.count(Registration.id)
    ).group_by(Registration.shift).all()

    return render_template('admin/reports/index.html',
                           bus_stats=bus_stats, shift_stats=shift_stats)
